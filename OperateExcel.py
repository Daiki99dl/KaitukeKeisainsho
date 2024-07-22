import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

def WriteDataFrameToExcel(df):
    # 必要に応じてCSVファイルに保存
    df.to_csv('output.csv', index=False, encoding='utf-8')

    # エクセルファイルの読み込み
    excel_path = r''
    book = load_workbook(excel_path)

    # コピーしたいシート名を指定
    sheet_name = 'テンプレート　コピー用'
    source_sheet = book[sheet_name]

    # 現在の日付を取得
    current_date = datetime.now().strftime('%Y%m%d')
    # コピーするシートの名前を指定
    new_sheet_name = f'自動転記_{current_date}'

    # シートのコピーを作成
    book.copy_worksheet(source_sheet).title = new_sheet_name
    
    # DataFrameをシートに書き込む（1行目にヘッダー、2行目からデータ）
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 1):
        for c_idx, value in enumerate(row, 1):
            book[new_sheet_name].cell(row=r_idx + 3, column=c_idx, value=value)    
    book.save(excel_path)
    book.close()
    print("DataFrameをExcelに書き込みました。")

